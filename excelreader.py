'''version 1.5
adding sheet names generator
return to excel function early if within A-Z range
adding generator'''

import time
import re
from collections import defaultdict
import openpyxl

def col_words(H):
    return ' '.join(x.strip() for x in re.findall(r'\D+',H))

def col_index(H):
    x = re.search(r'\d+', H)
    if x:
        return int(x.group(0))

def listc(x):
    ListSep = ','
    if type(x) is str:
        if ListSep in x:
            return [i for i in x.split(ListSep) if len(i)>0]
        else:
            return x


def to_excel(num):
    AZLIST = [chr(x + 65) for x in range(0,26)]
    ColsList = []
    while num > 0:
        mod = num % 26
        if mod == 0:
            num -= 1
        num = int(num / 26)
        ColsList.append(AZLIST[mod-1])
    return ''.join(reversed(ColsList))


def get_sheet_names(WorkBook):
    
    wbBuildData = openpyxl.load_workbook(WorkBook, data_only=True, read_only=True)
    ShNames = wbBuildData.get_sheet_names()
    Sheets = defaultdict(list)

    for SheetName in ShNames:
        yield SheetName

    print('Closing....')
    wbBuildData.close()

def get_wb(WorkBook):

    # Get the Sheet names in the workbook
    ShNames = list(get_sheet_names(WorkBook))
    Sheets = {}

    for SheetName in ShNames:
        print(SheetName)
        Sheets[SheetName] = list(get_sheet(WorkBook, SheetName))
    return Sheets


def get_sheet(WorkBook, SheetName, CheckInvalid=False, ChangeHeadings=False):
    
    wbBuildData = openpyxl.load_workbook(WorkBook, data_only=True, read_only=True)
    ShNames = wbBuildData.get_sheet_names()
    Sheets = defaultdict(list)

    Sh = wbBuildData.get_sheet_by_name(SheetName)

    Headings = [Sh[to_excel(x) + '1'].value for x in range(1,Sh.max_column+1)]


    for i in range(2,Sh.max_row+1):
        t0 = time.time()

        # Skip this row if we are checking for crossed out data
        if CheckInvalid:
            if any([Sh[to_excel(x+1) + str(i)].font.strikethrough for x in range(0, len(Headings)) if Sh[to_excel(x+1) + str(i)].value is not None]):
                continue

        Row = {Headings[x]:listc(Sh[to_excel(x+1) + str(i)].value) for x in range(0, len(Headings))}
        print(i, time.time() - t0)

        NewRow = defaultdict(dict)
        # Manipulate headings if desired
        if ChangeHeadings:
            # First merge columns with indexes.
            for K,V in Row.items():
                # If there is a number in the colum header
                if col_index(K):
                    NewRow[col_words(K)][col_index(K)] = V
                else:
                    NewRow[K] = V
        else:
            NewRow = Row

        yield NewRow
        

    wbBuildData.close()



if __name__=='__main__':
    x = get_sheet(r'test.xlsx', 'Sheet1')

    for Sh in x:
        print(Sh)


        

