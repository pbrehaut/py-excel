"""Version 9.4
Changing get_item_list to return empty list and get_item_indicies to return None upon key error
Need to fix column double letters issue.
changing the swap header to be case insensitive
added exception handling to get_val to catch key errors and return a text error, means no need for error handling to catch key errors in main code
added some error handling and validation and make sheet names case not case sensitive
attempt to open version sheet in init
Modify cursor methods to skip over rows not in selection
split out repeated code into functions for build_dict and build_list_dict_row
Adding a version functionality to select rows to be gathered.
Class ReadSheet - parent class use a row index when retrieving data via methods
Class ReadShseetCursor -sub class of ReadSheet which overwrites methods to not require a row index
and use a move next method to travserse rows
Class F5ReadSheet/NXReadSheet - overwrites the function to swap names with the
translations specified in the object dictionary - inherit from either above objects"""

class ReadSheet:
    def __init__(self, WorkBook, Sheet, VersionSheetName="version", VersionNumber=0):
        """Open excel sheets and build the dictionary to use"""
        import openpyxl
        self.wbBuildData = openpyxl.load_workbook(WorkBook, data_only=True, read_only=True)

        #  Lower the case for comparison of the sheet names to the sheet name argument given.
        for x in self.wbBuildData.get_sheet_names():
            if x.lower() == Sheet.lower():
                Sheet=x
                
        self.shBuildData = self.wbBuildData.get_sheet_by_name(Sheet)
        #  Dictionary to put excel data into.
        self.BuildDict = {}
        #  Get the build data info from the sheet.
        self.BuildDict = self.build_dict(self.shBuildData)

        # Find the version control sheet and call the functions to find the rows selected for the version.
        # If no version number is given use the highest found.
        # If no version sheet is given use a sheet named 'version' if it exists in the excel workbook.
        # If no version sheet is found or given then process all the rows.
        # If no rows are specified for the selected version raise an error message and end.
        
        # Lower the case for comparison of the version sheet name given with the sheets in this workbook.
        for x in self.wbBuildData.get_sheet_names():
            if x.lower() == VersionSheetName.lower():
                VersionSheetName=x
                self.shVersionSheet = self.wbBuildData.get_sheet_by_name(VersionSheetName)
                self.version_control(VersionSheetName, VersionNumber)
                # If version control is being use we need to set this to lowest row in the version row list
                # this will start the rows at the desired first row.
                self.CurrentRow = self.VersionRows[0]

                # If the version sheet is found break out of loop otherwise it will loop through again
                # and go to the else statement below and disable version control.
                break
            else:
                # Intialise the row pointer to the first row if there is no version tab found.
                # We will check later if VersionRows is set to None to decide whether to skip over rows.
                self.CurrentRow = 1
                self.VersionRows = None
        
    def build_dict(self, Sheet):
        """Returns a dictionary from the sheet supplied"""
        # If these are found in an excel data cell then assume it is a list of values and split them into a list.
        ListDelimiters=[',', ';']
        BuildDict = {}
        ColumnDict = self.get_cols(Sheet)
        
        # Go through each row.
        RowCount=Sheet.max_row+1
        for i in range(2, RowCount):
            BuildDict[i-1] = {}
        
            # Add nested dictionary for each header so we can add the index.
            for j in ColumnDict:
                x = self.get_field_spec(j, ColumnDict[j])
                # Replace the header names with the values we want for F5 fields with the 'header_to_name method.
                BuildDict[i-1][self.header_to_name(x['fieldName'])] = {}

            # Add the actual values to the nested dictionary.
            for j in ColumnDict:
                x = self.get_field_spec(j, ColumnDict[j])
                ValueAdd = Sheet[x['letter'] + str(i)].value

                # If there is a field separator then split the cell data into a list and add that to the dictionary.
                for y in ListDelimiters:
                    if y in str(ValueAdd):

                        # If there is an index specified other than 1 with the column header for a list raise an error and exit
                        # since the methods for lists won't work with multiple columns.
                        if not x['itemIndex'] == 1:
                            ListErrStr = "Cannot use an index other than 1 for lists. Remove the number "
                            ListErrStr += str(x['itemIndex']) + " from " + x['fieldName'] + " in column "
                            ListErrStr += ColumnDict[j] + " or set it to 1"
                            raise Exception(ListErrStr)
                        else:
                            ValueAdd = ValueAdd.split(y)
                            BuildDict[i-1][self.header_to_name(x['fieldName'])]['type'] = 'list'
                            break
                BuildDict[i-1][self.header_to_name(x['fieldName'])][x['itemIndex']] = ValueAdd
        return BuildDict

    def version_control(self, VersionSheet, VersionNumber):
        """Invoke this method to enable record selection based on a version number to row mapping"""
        VersionDict = {}
        VersionDict = self.build_dict(self.shVersionSheet)
        self.VersionLists = {}

        try:
            for i in VersionDict:
                Version = VersionDict[i]['version'][1]
                self.VersionLists[Version] = []
        except KeyError:
            raise Exception("Version column not found in sheet: " + VersionSheet + ". Please add version column or remove the version sheet or don't specify a version sheet")
            
        for i in VersionDict:
            Version = VersionDict[i]['version'][1]

            # If the data is already a list then just add it using +=.
            # Otherwise append the values to the list for the version number.
            try:
                if 'type' in VersionDict[i]['rows']:
                    if VersionDict[i]['rows']['type'] == 'list':
                        self.VersionLists[Version]+=VersionDict[i]['rows'][1]
                    else:
                        self.VersionLists[Version].append(VersionDict[i]['rows'][1])
                else:
                    self.VersionLists[Version].append(VersionDict[i]['rows'][1])
            except KeyError:
                raise Exception("Rows column not found in sheet: " + VersionSheet + ". Please add rows column or remove the version sheet or don't specify a version sheet")
                
        # Now convert all row values to a list of row numbers.
        # E.g convert 3-6 to 3,4,5,6
        for i in self.VersionLists:
            # Don't call the function if there are no rows specified in the row column.
            if None not in self.VersionLists[i]:
                self.VersionLists[i] = self.range_to_list(self.VersionLists[i])

        # Populate attribute VersionRows with the list of rows to select for the given version number or the highest
        # version found in the version sheet.
        self.VersionRows=self.get_version_rows(VersionNumber)

        # Subtract 1 from each row number that the user supplied as they would be excel row numbers
        # not dictionary row numbers. The Row numbers in excel should start at 2 instead of 1 for our dictionary.
        for i in range(0,len(self.VersionRows)):
            self.VersionRows[i] -= 1
        
    def get_version_rows(self,VersionNumber):
        """Return a list of rows for the given version or if none specified use the highest version found"""
        x = VersionNumber
        if VersionNumber == 0:
            x = VersionNumber
            
            # Find the highest index number if no version argument is given.
            for i in self.VersionLists:
                if i > x:
                    x = i
                    
        # If the selected version has no rows raise an error.
        if None in self.VersionLists[x]:
            raise Exception("Selected version " + str(x) + " has no rows, select a different version or add a row selection to version " + str(x))
        else:
            return sorted(self.VersionLists[x])            

    def range_to_list(self,RangeList):
        """Return a list replacing ranges with each number in the range"""
        NewList = []
        for i in RangeList:
            if "-" in str(i):
                a = i.split("-")
                for x in range(int(a[0]), int(a[1]) + 1):
                    NewList.append(x)
            else:
                NewList.append(int(i))
        return NewList                        

    def get_cols(self,Sheet):
        """Return a dictionary of the column name and letter reference"""
        import string
        ColumnCount = Sheet.max_column
        if ColumnCount <= 26:
            Columns = dict(zip(range(0, ColumnCount), string.ascii_uppercase))
        else:
            Columns = dict(zip(range(0, 26), string.ascii_uppercase))
            for x in range(26, ColumnCount):
                Columns[x] = 'A' + Columns[x-26]
            
        ColumnHeaders = {}
        for ColNum in Columns:
            ColLetter = Columns[ColNum]
            ColValue = Sheet[ColLetter + str(1)].value
            ColumnHeaders[ColValue] = ColLetter
        return ColumnHeaders

    def get_field_spec(self, ColHeader, ColLetter):
        """Return a dictionary of the field name and index gleaned from the header and the excel column letter e.g {fieldName:node,index:2,letter:'C'}"""
        FieldSpec = {}

        # Find the digit in the header and put it in the dictionary.
        # Put the remaining words together.
        x = ColHeader.split()
        Header = []
        ItemIndex = 1
        for i in x:
            if i.isdigit():
                ItemIndex = i
            else:
                Header.append(i)
        Header=" ".join(Header)
            
        FieldSpec['itemIndex'] = int(ItemIndex)
        FieldSpec['fieldName'] = Header
        FieldSpec['letter'] = ColLetter       
        return FieldSpec
    
    def header_to_name(self, Header):
        """Parent class method to just return the header value back, this method will be overwritten for sub-classes"""
        return self.swap_header(Header)

    def swap_header(self, Header):
        """Return the standard header name from our dictionary, or just the given header value back if there is no match"""
        # Find the Header value supplied in our dictionary of terms and return the standard name we should use for fields.
        try:
            for i in self.ColumnTr:
                for j in self.ColumnTr[i]:
                    if Header.lower() == j.lower():
                        return i
            else:
                return Header.lower()
        except AttributeError:
            return Header.lower()

    # ------------------------------------------------------
    # Methods for data access from the object
    # ------------------------------------------------------

    def get_row_count(self, Sheet):
        """Return the count of rows"""
        rows=Sheet.max_row
        return rows

    def get_val(self,row,field,itemIndex=1):
        """Return the value at the given row, field and index. If field not there return None"""
        try:
            ReturnVal = self.BuildDict[row][field][itemIndex]
            # Catch any key errors here so the calling script won't need to catch key errors and can just check for the None value.
            # This allows calling scripts to just test for None which covers the field not being there and also covers no data in the field.
        except KeyError:
            ReturnVal = None
        return ReturnVal

    def get_count(self,row,field):
        """Return the count of columns there are for the given field, not the count of actual values there"""
        return len(self.BuildDict[row][field])

    def get_rows(self):
        """Return a list of row indicies"""
        a = []
        for i in self.BuildDict:
            a.append(i)
        return a

    def get_item_list(self,row,field):
        """Return a list of actual values stored for this field"""
        try:
            x = []
            # If the type key is set to list then don't build up a list just return what is there for the first index.
            # This means the user can't put in an index in the column header when putting in a list.
            if 'type' in self.BuildDict[row][field]:
                if not self.BuildDict[row][field]['type'] == 'list':
                    x = self.build_list_dict_row(self.BuildDict[row][field])
                else:
                    return self.BuildDict[row][field][1]
            else:
                # If the type key isn't there at all then assume we need to build a standard list.
                x = self.build_list_dict_row(self.BuildDict[row][field])
        except KeyError:
            x = []
        return x
        

    def get_item_indicies(self,row,field):
        """Return a list of the indicies that have actual values for this field"""
        # This method used along with get_val will be one of the main ways to extract data as it provides a list of indicies to loop through.
        # The indicies returned will be used with get_val.
        try:
            x = []
            for i in self.BuildDict[row][field]:
                if self.BuildDict[row][field][i] is not None:
                    x.append(i)
        except KeyError:
            x = None
        return x

    def build_list_dict_row(self,DictRow):
        """Return a list based on joining up the fields for a given header and different sub-indicies"""
        # This method will only be useful if the user has put in standard well ordered columns and values.
        # This could also be used if you do not need to match up indicies from different columns.
        # If you need to match up 'port 1' with 'port 1 config' it could be problematic.
        x = []
        for i in DictRow:
            if DictRow[i] is not None and str(i).isdigit():
                x.append(DictRow[i])
        return x

class ReadSheetCursor(ReadSheet):
    # Use this subclass instead of ReadSheet if you want to use functions without having to specify the row index.
    # You need to iterate over rows using methods.
    # The over written functions here basically set the row to the current row instead of relying on a row index supplied.
    # The move_next method increments CurrentRow to go to the next record.
    
    def move_next(self):
        """Move to the next row"""
        # This method increments the CurrentRow counter.
        # This counter is then used in other methods for specifiying the row index.
        self.CurrentRow += 1
        if self.VersionRows is not None:
            while self.end_of_rows() is False:
                # If the current row is in the filter/version row list break out.
                if self.CurrentRow in self.VersionRows:
                    break
                else:
                    # Keep looping through row numbers till we reach the end of rows or finding a matching row in the filter.
                    self.CurrentRow += 1
                
    def end_of_rows(self):
        """Returns false until we are past the last row"""
        if self.CurrentRow == self.get_row_count(self.shBuildData):
            return True
        else:
            return False
        
    def get_item_list(self,field):
        """Return a list of actual values stored for this field"""
        row = self.CurrentRow
        try:
            x = []
            # If the type key is set to list then don't build up a list just return what is there for the first index.
            # This means the user can't put in index in the column header when putting in a list.
            if 'type' in self.BuildDict[row][field]:
                if not self.BuildDict[row][field]['type'] == 'list':
                    x = self.build_list_dict_row(self.BuildDict[row][field])
                else:
                    return self.BuildDict[row][field][1]
            else:
                # If the type key isn't there at all then assume we need to build a standard list.
                x = self.build_list_dict_row(self.BuildDict[row][field])
        except KeyError:
            x = []
        return x
    
    def get_val(self,field,itemIndex=1):
        """Return the value at the given row, field and index"""
        row = self.CurrentRow
        try:
            ReturnVal = self.BuildDict[row][field][itemIndex]
        except KeyError:
            ReturnVal = None
        return ReturnVal

    def get_item_indicies(self,field):
        """Return a list of the indicies that have actual values for this field"""
        row = self.CurrentRow
        try:
            x = []
            for i in self.BuildDict[row][field]:
                if self.BuildDict[row][field][i] is not None and str(i).isdigit():
                    x.append(i)
        except KeyError:
            x = None
        return x

class NXReadSheet(ReadSheetCursor):
    def header_to_name(self, Header):
        """subclass method to return the standard F5 field name by looking up the dictionary defined within this method"""
        self.ColumnTr = {}
        self.ColumnTr['device'] = ["switch", "host", "hostname", "router"]
        self.ColumnTr['ip'] = ["ip address"]
        return self.swap_header(Header)
