import openpyxl
from openpyxl.styles import PatternFill, Font

MAXWIDTH = 200
PADDING = 2

def get_lol_len(DataStruct):
    '''return the number of fields in a dict'''
    x=0
    for i in DataStruct:
        if len(i) > x:
            x = len(i)
    return x

def normalise(DataStruct, Headings):
    '''if the data is just a dict then flatten it to a list of lists'''
    def flatten_dict(D):
        def iteritems_recursive(d):
          for k,v in d.items():
            if isinstance(v, dict):
              for k1,v1 in iteritems_recursive(v):
                yield [k,]+k1, v1
            else:
              yield [k,],v

        NewList = []
        for p,v in iteritems_recursive(D):
            p.append(v)
            NewList.append(p)
        return NewList
    
    if get_data_type(DataStruct) == 'Dict':
        x = flatten_dict(DataStruct)
        
        if len(Headings) == 0:
            DummyHeadings = []
            # Insert dummy headings into the first line of the datastruct
            for i in range(1,(get_lol_len(x)+1)):
                DummyHeadings.append('Col' + str(i))
            x.insert(0,DummyHeadings)
        return x
    else:
        return DataStruct


def get_data_type(DataStruct):
        if type(DataStruct) is list and type(DataStruct[0]) is list:
                return 'ListofLists'
        elif type(DataStruct) is list and type(DataStruct[0]) is dict:
                return 'ListofDicts'
        elif type(DataStruct) is dict:
                return 'Dict'


def get_XLS_filename(xlsReportName):
        # Check the report name has the correct ext
        if '.' in xlsReportName:
            if xlsReportName.split('.')[-1] != '.xlsx':
                FileSplit = xlsReportName.split('.')[:-1]
                FileSplit.append('xlsx')
                xlsReportName = '.'.join(FileSplit)
        else:
            xlsReportName += '.xlsx'
        return xlsReportName


def to_excel(num):
    AZLIST = [chr(x + 65) for x in range(0,26)]
    ColsList = []
    while num > 0:
        mod = num % 26
        if mod == 0:
            num -= 1
        num = int(num / 26)
        ColsList.append(AZLIST[mod-1])
    return ''.join(reversed(ColsList))


def get_headings(DataStruct, Headings):
        # If column headings were provided use them.
        if len(Headings) > 0:
            Headings = Headings

        # List of lists:
        elif get_data_type(DataStruct) == 'ListofLists':
            # Return the first row then delete it.
            Headings = DataStruct[0]
            del(DataStruct[0])

        # List of dicts:
        elif get_data_type(DataStruct) == 'ListofDicts':
            # Go through all lines and get all Keys to use as headings.
            for Line in DataStruct:
                for K in Line:
                    if K not in Headings:
                        Headings.append(K)

        elif get_data_type(DataStruct) == 'Dict':
            Headings = [x for x in DataStruct]

        return Headings


def get_colspec(DataStruct, Headings=None):

    if Headings is None:
        Headings = []
        
    Headings = get_headings(DataStruct, Headings)
    ColSpec = {}
    
    # If we are writing a list to excel.
    if get_data_type(DataStruct) == 'ListofLists':
        
        # Add the headings into the Colspec dict
        for i, Heading in enumerate(Headings):
            Col = to_excel(i+1)    
            ColSpec[Col] = {}
            ColSpec[Col]['Width'] = len(Heading)
            ColSpec[Col]['Heading'] = Heading
                
        # Go through each line.
        for Line in DataStruct:
                # Go through each item in the line.
                for i, Item in enumerate(Line):
                    Col = to_excel(i+1)
                    if Col not in ColSpec:
                        ColSpec[Col] = {}
                        ColSpec[Col]['Width'] = 0
                    if len(str(Item)) > ColSpec[Col]['Width']:
                        ColSpec[Col]['Width'] = len(Item)

    # If we are writing a list of dicts to excel.
    elif get_data_type(DataStruct) == 'ListofDicts':
        # Go through each item in the Headings list
        for i, H in enumerate(Headings):
            Col = to_excel(i+1)
            ColSpec[Col] = {}
            ColSpec[Col]['Width'] = len(H)
            ColSpec[Col]['Heading'] = H
            # Loop through the data to get teh column width
            for Line in DataStruct:
                    for K, V in Line.items():
                            if H in Line and H == K:
                                    if len(str(V)) > ColSpec[Col]['Width'] < MAXWIDTH:
                                            ColSpec[Col]['Width'] = len(str(V))

    # Add padding to the column width
    # Find any columns with width > MAXWIDTH and change them
    for Col in ColSpec:
        ColSpec[Col]['Width'] += PADDING
        if ColSpec[Col]['Width'] > MAXWIDTH:
            ColSpec[Col]['Width'] = MAXWIDTH

    return ColSpec


def get_col_from_h(ColSpec, Heading):
        '''find the heading and return the column'''
        for K, V in ColSpec.items():
                if V['Heading'] == Heading:
                        return K

def write_sheet(FileName, DataStruct, SheetName='Sheet1', Headings=None):

    # Make Headings an empty list if none are passed in
    if Headings == None:
        Headings = []

    GRAYFILL = PatternFill(start_color='808080', end_color='808080',fill_type='solid')

    wb = openpyxl.Workbook()
    wb.create_sheet(index=0, title=SheetName)
    Sheet = wb.get_sheet_by_name(SheetName)

    DataStruct = normalise(DataStruct, Headings)
    
    ColSpec = get_colspec(DataStruct, Headings)
    #print(ColSpec)
    # Write headings and set column widths.
    for K, V in ColSpec.items():
        Sheet.column_dimensions[K].width = V['Width']
        Sheet[K + '1'] = V['Heading']
        Sheet[K + '1'].fill = GRAYFILL
        Sheet[K + '1'].font = Font(bold=True)

    # Write the rest of the data.
    # If list of lists.
    if get_data_type(DataStruct) == 'ListofLists':
            for NL, Line in enumerate(DataStruct):
                # Start at 2 instead of 0, we have already printed the column headers.
                NL += 2
                for NF, Item in enumerate(Line):
                    Cell = to_excel(NF + 1) + str(NL)
                    Sheet[Cell] = Item
    elif get_data_type(DataStruct) == 'ListofDicts':
            for NL, Line in enumerate(DataStruct):
                NL += 2
                for K, Item in Line.items():
                        Cell = get_col_from_h(ColSpec, K) + str(NL)
                        try:
                            Sheet[Cell] = Item
                        except:
                            print(Item)
                            Sheet[Cell] = 'Illegal Character Error'
    wb.save(get_XLS_filename(FileName))
    wb.close()


if __name__=='__main__':
    # Create test data
    # List of lists
    Report = []
    for rown in range(0,10):
        Report.append([f'Row {rown} Col {x}' * 3 for x in range(0,5)])

    write_sheet('test_lists.xlsx', Report, SheetName='Lists')

    Headings = [f'Col {x} Heading' for x in range(0,5)]
    write_sheet('test_lists_w_headings.xlsx', Report, SheetName='Lists', Headings=Headings)

    # List of Dicts
    Report = []
    for rown in range(0,10):
        Report.append({f'Col {x} Heading':'Dummy value' for x in range(0,5)})

    write_sheet('test_dicts.xlsx', Report, SheetName='Dicts')


    




